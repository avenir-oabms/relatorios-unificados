from flask import Blueprint, request, jsonify, send_file
from functools import wraps
from sqlalchemy import text
from db import MySQLSession, MSSQLSession, ping_mysql, ping_mssql
from auth import verify_token

# ===== imports para geração de arquivos =====
import io, os, zipfile, datetime as dt
import pandas as pd
import tempfile
import traceback

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
)
from reportlab.lib.units import mm

bp = Blueprint("reports", __name__, url_prefix="/api/reports")


# -------------------------------------------------------
#                SAÚDE DE BANCO / AUTH
# -------------------------------------------------------
@bp.get("/health/db")
def health_db():
    """
    Diagnóstico das conexões de banco.
    - mysql: { ok: bool, value?: 1, error?: str }
    - mssql: { ok: bool, value?: 1, error?: str }
    """
    return jsonify({
        "mysql": ping_mysql(),
        "mssql": ping_mssql(),
    })


def require_auth(f):
    @wraps(f)
    def _wrap(*args, **kwargs):
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        u = verify_token(token)
        if not u:
            return jsonify({"error": "Não autorizado"}), 401
        request.user = u
        return f(*args, **kwargs)
    return _wrap


def user_has_report(uid: int, report_key: str) -> bool:
    with MySQLSession() as s:
        row = s.execute(text("""
            SELECT 1
            FROM reports r
            JOIN report_permissions rp ON rp.report_id = r.id
            WHERE rp.user_id = :uid AND r.report_key = :rk
            LIMIT 1
        """), {"uid": uid, "rk": report_key}).first()
        return row is not None


# -------------------------------------------------------
#                     LISTAGEM DE RELATÓRIOS
# -------------------------------------------------------
@bp.get("/list")
@require_auth
def list_reports():
    uid = request.user["uid"]
    with MySQLSession() as s:
        rows = s.execute(text("""
            SELECT r.report_key AS `key`, r.module, r.label
            FROM reports r
            JOIN report_permissions rp ON rp.report_id = r.id
            WHERE rp.user_id = :uid
            ORDER BY r.module, r.label
        """), {"uid": uid}).mappings().all()

    grouped = {}
    for r in rows:
        grouped.setdefault(r["module"], []).append({"key": r["key"], "label": r["label"]})
    return jsonify(grouped)


# -------------------------------------------------------
#                     EXEMPLOS DE RUN
# -------------------------------------------------------
@bp.post("/run/<report_key>")
@require_auth
def run_report(report_key):
    uid = request.user["uid"]
    if not user_has_report(uid, report_key):
        return jsonify({"error": "Sem permissão"}), 403

    if report_key == "adm_usuarios":
        with MySQLSession() as s:
            rows = s.execute(text("""
                SELECT id, name AS Nome, email AS Email, active AS Ativo, created_at AS CriadoEm
                FROM users
                ORDER BY id
            """)).mappings().all()
        cols = list(rows[0].keys()) if rows else []
        return jsonify({"columns": cols, "rows": rows, "total_rows": len(rows)})

    if report_key == "fin_inadimplencia_resumo":
        with MSSQLSession() as s:
            rows = s.execute(text("""
                SELECT TOP 100
                    suc.NomeSubUnidade AS Subsecao,
                    COUNT(p.ID) AS TotalInscritos
                FROM Pessoa p
                LEFT JOIN SubUnidadeConselho suc ON p.SubUnidadeAtual = suc.ID
                WHERE p.TipoCategoria = 20
                GROUP BY suc.NomeSubUnidade
                ORDER BY suc.NomeSubUnidade
            """)).mappings().all()
        cols = list(rows[0].keys()) if rows else []
        return jsonify({"columns": cols, "rows": rows, "total_rows": len(rows)})

    return jsonify({"error": "Relatório desconhecido"}), 404


# -------------------------------------------------------
#                RELATÓRIO: LISTA SIMPLES
# -------------------------------------------------------
def _consulta_lista_simples(subsecao_like: str | None) -> pd.DataFrame:
    """Consulta base (MSSQL) com filtro opcional de subseção."""
    try:
        filtro = ""
        params = {}
        if subsecao_like:
            filtro = "AND suc.NomeSubUnidade LIKE :sub"
            params["sub"] = f"%{subsecao_like}%"

        sql = f"""
            SELECT 
                p.RegistroConselhoAtual AS OAB, 
                p.Nome,
                p.CPFCNPJ,
                s.Descricao AS Situacao,
                FORMAT(p.DataNascimentoFundacao, 'dd/MM/yyyy') AS DataNascimento,
                FORMAT(p.DataCompromisso, 'dd/MM/yyyy') AS DataCompromisso,
                p.TelefoneCelular, 
                COALESCE(p.EmailCorreio, p.EmailComercial) AS Email,
                suc.NomeSubUnidade AS Subsecao
            FROM Pessoa p
            JOIN SubUnidadeConselho suc ON p.SubUnidadeAtual = suc.ID
            JOIN Situacao s ON p.SituacaoAtual = s.ID
            WHERE p.TipoCategoria = 20
              AND p.SituacaoAtual = 14
              {filtro}
            ORDER BY p.Nome
        """
        
        with MSSQLSession() as s:
            rows = s.execute(text(sql), params).mappings().all()
        
        # Converte para DataFrame garantindo que todos os valores são strings ou None
        data = []
        for row in rows:
            row_dict = {}
            for key, value in row.items():
                if value is None:
                    row_dict[key] = ""
                else:
                    row_dict[key] = str(value)
            data.append(row_dict)
        
        return pd.DataFrame(data)
        
    except Exception as e:
        print(f"Erro na consulta: {e}")
        traceback.print_exc()
        return pd.DataFrame()


# ---------- PDF: Com suporte a orientação retrato/paisagem ----------
def _pdf_from_df(df: pd.DataFrame, titulo: str, subsecao: str, campos_selecionados: list = None, orientacao: str = "paisagem") -> io.BytesIO:
    buf = io.BytesIO()

    # Definir orientação da página baseada no parâmetro
    if orientacao == "retrato":
        page_size = A4  # A4 retrato
        left_margin = right_margin = 20 * mm
        top_margin = 30 * mm
        bottom_margin = 25 * mm
    else:
        page_size = landscape(A4)  # A4 paisagem (padrão)
        left_margin = right_margin = 12 * mm
        top_margin = 24 * mm
        bottom_margin = 20 * mm

    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Tiny", fontSize=8, leading=10))
    styles.add(ParagraphStyle(name="TitleCenter", parent=styles["Heading1"], alignment=1))

    # tenta 2 nomes de arquivo para a logo
    base_dir = os.path.dirname(__file__)
    logo_candidates = [
        os.path.join(base_dir, "static", "logos", "logo_oabms.png"),
        os.path.join(base_dir, "static", "logos", "logo-oab.png"),
        os.path.join(base_dir, "static", "logo_oabms.png"),
        os.path.join(base_dir, "static", "logo-oab.png"),
    ]
    logo_path = next((p for p in logo_candidates if os.path.exists(p)), None)

    def _header_footer(canv, _doc):
        canv.saveState()
        W, H = page_size

        # --- Cabeçalho ---
        # Logo à esquerda
        if logo_path:
            try:
                img_w, img_h = 28*mm, 14*mm
                canv.drawImage(
                    logo_path,
                    left_margin,
                    H - img_h - 6*mm,
                    width=img_w,
                    height=img_h,
                    preserveAspectRatio=True,
                    mask='auto'
                )
            except Exception:
                pass

        # Título centralizado
        canv.setFont("Helvetica-Bold", 14)
        canv.drawCentredString(W/2, H - 12*mm, titulo)

        # Texto à direita: Subseção + data
        info = f"Subseção: {subsecao or 'Geral'}  |  Gerado em: {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}"
        canv.setFont("Helvetica", 9)
        canv.drawRightString(W - right_margin, H - 17*mm, info)

        # --- Rodapé (LGPD) em TODAS as páginas ---
        lgpd_txt = (
            "LGPD – Aviso: Este relatório contém dados pessoais destinados exclusivamente às atividades "
            "institucionais da OAB/MS. É vedado o uso para fins distintos, devendo o destinatário adotar "
            "medidas de segurança compatíveis com a Lei nº 13.709/2018."
        )
        p = Paragraph(lgpd_txt, styles["Tiny"])
        avail_w = W - left_margin - right_margin
        w, h = p.wrap(avail_w, 30*mm)
        p.drawOn(canv, left_margin, bottom_margin - h + 2)

        canv.restoreState()

    # -------- Story (tabela) --------
    story = []

    if df is None or df.empty:
        story.append(Paragraph("Nenhum registro encontrado.", styles["Normal"]))
        doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
        buf.seek(0)
        return buf

    # NOVO: Configurar colunas e larguras dinamicamente baseado nos campos selecionados
    if campos_selecionados:
        # Configurações diferentes para retrato vs paisagem
        if orientacao == "retrato":
            campo_config = {
                "OAB": {"col": "OAB", "width": 20*mm, "display": "OAB"},
                "Nome": {"col": "Nome", "width": 80*mm, "display": "Nome"},
                "CPF/CNPJ": {"col": "CPFCNPJ", "width": 30*mm, "display": "CPF/CNPJ"},
                "Situacao": {"col": "Situacao", "width": 25*mm, "display": "Situação"},
                "DataNascimento": {"col": "DataNascimento", "width": 28*mm, "display": "Data Nasc."},
                "DataCompromisso": {"col": "DataCompromisso", "width": 28*mm, "display": "Compromisso"},
                "TelefoneCelular": {"col": "TelefoneCelular", "width": 32*mm, "display": "Celular"},
                "Email": {"col": "Email", "width": 65*mm, "display": "E-mail"},
                "Subsecao": {"col": "Subsecao", "width": 35*mm, "display": "Subseção"}
            }
        else:  # paisagem
            campo_config = {
                "OAB": {"col": "OAB", "width": 16*mm, "display": "OAB"},
                "Nome": {"col": "Nome", "width": 65*mm, "display": "Nome"},
                "CPF/CNPJ": {"col": "CPFCNPJ", "width": 27*mm, "display": "CPF/CNPJ"},
                "Situacao": {"col": "Situacao", "width": 20*mm, "display": "Situação"},
                "DataNascimento": {"col": "DataNascimento", "width": 22*mm, "display": "Data Nasc."},
                "DataCompromisso": {"col": "DataCompromisso", "width": 25*mm, "display": "Compromisso"},
                "TelefoneCelular": {"col": "TelefoneCelular", "width": 25*mm, "display": "Celular"},
                "Email": {"col": "Email", "width": 62*mm, "display": "E-mail"},
                "Subsecao": {"col": "Subsecao", "width": 27*mm, "display": "Subseção"}
            }
        
        # Construir listas de colunas e larguras apenas para campos selecionados
        columns = []
        col_widths = []
        df_columns = []
        
        for campo in campos_selecionados:
            if campo in campo_config and campo_config[campo]["col"] in df.columns:
                config = campo_config[campo]
                columns.append(config["display"])
                col_widths.append(config["width"])
                df_columns.append(config["col"])
        
        # Filtrar DataFrame para incluir apenas as colunas selecionadas
        if df_columns:
            df = df[df_columns]
            
        # Renomear colunas para display
        rename_dict = {}
        for campo in campos_selecionados:
            if campo in campo_config and campo_config[campo]["col"] in df.columns:
                old_name = campo_config[campo]["col"]
                new_name = campo_config[campo]["display"]
                rename_dict[old_name] = new_name
        
        if rename_dict:
            df = df.rename(columns=rename_dict)
    else:
        # Configuração padrão - ajustar baseado na orientação
        df = df.rename(columns={
            "Situacao": "Situação",
            "DataNascimento": "Data Nasc.",
            "DataCompromisso": "Compromisso",
            "TelefoneCelular": "Celular",
            "Subsecao": "Subseção",
            "CPFCNPJ": "CPF/CNPJ"
        })
        
        columns = ["OAB","Nome","CPF/CNPJ","Situação","Data Nasc.","Compromisso","Celular","E-mail","Subseção"]
        
        if orientacao == "retrato":
            col_widths = [20*mm, 80*mm, 30*mm, 25*mm, 28*mm, 28*mm, 32*mm, 65*mm, 35*mm]
        else:  # paisagem
            col_widths = [16*mm, 65*mm, 27*mm, 20*mm, 22*mm, 25*mm, 25*mm, 62*mm, 27*mm]

    def P(text):
        return Paragraph((str(text) if text is not None else "").replace("&", "&amp;"), styles["Tiny"])

    data = [columns]
    for _, r in df.iterrows():
        row_data = []
        for col in df.columns:
            row_data.append(P(r.get(col, "")))
        data.append(row_data)

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#23364B")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("ALIGN",      (0,0), (-1,0), "CENTER"),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 9),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),

        ("FONTNAME", (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,1), (-1,-1), 8),
        ("VALIGN",   (0,1), (-1,-1), "TOP"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor("#F7F9FC")]),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#B6C2CF")),
    ]))

    story.append(tbl)
    story.append(Spacer(1, 6))  # respiro final

    doc.build(story, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buf.seek(0)
    return buf


def _excel_from_df(df: pd.DataFrame, titulo: str, subsecao: str, campos_selecionados: list = None) -> io.BytesIO:
    """Gera arquivo Excel com formatação melhorada."""
    try:
        bio = io.BytesIO()
        
        # Prepara os dados
        if df.empty:
            df_export = pd.DataFrame({"Mensagem": ["Nenhum registro encontrado"]})
        else:
            df_export = df.copy()
            # Renomeia colunas para português apenas se elas existirem
            rename_map = {
                "Situacao": "Situação",
                "DataNascimento": "Data Nascimento",
                "DataCompromisso": "Data Compromisso",
                "TelefoneCelular": "Telefone Celular",
                "Subsecao": "Subseção",
                "CPFCNPJ": "CPF/CNPJ"
            }
            # Aplicar apenas as renomeações para colunas que existem
            rename_dict = {k: v for k, v in rename_map.items() if k in df_export.columns}
            if rename_dict:
                df_export = df_export.rename(columns=rename_dict)
        
        # Cria o arquivo Excel
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            # Escreve os dados
            df_export.to_excel(writer, index=False, sheet_name="Lista de Inscritos", startrow=3)
            
            # Acessa o workbook e worksheet
            workbook = writer.book
            worksheet = writer.sheets["Lista de Inscritos"]
            
            # Adiciona cabeçalho
            worksheet["A1"] = titulo
            worksheet["A2"] = f"Subseção: {subsecao or 'Geral'} | Gerado em: {dt.datetime.now().strftime('%d/%m/%Y %H:%M')}"
            
            # Formatação do cabeçalho
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            title_font = Font(bold=True, size=14)
            subtitle_font = Font(size=10)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="23364B", end_color="23364B", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplica formatação ao título
            worksheet["A1"].font = title_font
            worksheet["A2"].font = subtitle_font
            
            # Formatação dos cabeçalhos da tabela (linha 4)
            if not df.empty:
                for col_num, column_title in enumerate(df_export.columns, 1):
                    cell = worksheet.cell(row=4, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                
                # Aplica bordas aos dados
                for row in range(4, len(df_export) + 5):
                    for col in range(1, len(df_export.columns) + 1):
                        worksheet.cell(row=row, column=col).border = thin_border
                
                # Ajusta largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        bio.seek(0)
        return bio
        
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        traceback.print_exc()
        # Fallback simples
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            (df or pd.DataFrame()).to_excel(writer, index=False, sheet_name="Lista")
        bio.seek(0)
        return bio


def _csv_from_df(df: pd.DataFrame, titulo: str, subsecao: str, campos_selecionados: list = None) -> io.BytesIO:
    """Gera arquivo CSV limpo para importações - apenas dados, sem cabeçalhos informativos."""
    try:
        bio = io.BytesIO()
        
        if df.empty:
            # CSV vazio - apenas uma linha indicativa
            csv_content = "Nenhum registro encontrado\n"
            bio.write(csv_content.encode('utf-8-sig'))
        else:
            # Renomeia colunas para português apenas se elas existirem
            rename_map = {
                "Situacao": "Situação",
                "DataNascimento": "Data Nascimento", 
                "DataCompromisso": "Data Compromisso",
                "TelefoneCelular": "Telefone Celular",
                "Subsecao": "Subseção",
                "CPFCNPJ": "CPF/CNPJ"
            }
            # Aplicar apenas as renomeações para colunas que existem
            rename_dict = {k: v for k, v in rename_map.items() if k in df.columns}
            if rename_dict:
                df_export = df.rename(columns=rename_dict)
            else:
                df_export = df.copy()
            
            # Gera CSV limpo - apenas cabeçalho das colunas + dados
            df_export.to_csv(bio, index=False, sep=";", encoding="utf-8-sig")
        
        bio.seek(0)
        return bio
        
    except Exception as e:
        print(f"Erro ao gerar CSV: {e}")
        traceback.print_exc()
        # Fallback simples
        bio = io.BytesIO()
        (df or pd.DataFrame()).to_csv(bio, index=False, sep=";", encoding="utf-8-sig")
        bio.seek(0)
        return bio


# -------------------------------------------------------
#                SUPORTE: SUBSEÇÕES PARA UI
# -------------------------------------------------------
@bp.get("/subsecoes")
@require_auth
def subsecoes():
    """
    Endpoint para buscar lista de subseções disponíveis.
    Retorna apenas subseções com TipoSubUnidade = 2 (subseções ativas).
    """
    try:
        with MSSQLSession() as s:
            rows = s.execute(text("""
                SELECT ID, NomeSubUnidade
                FROM SubUnidadeConselho
                WHERE TipoSubUnidade = 2
                ORDER BY NomeSubUnidade
            """)).mappings().all()
        
        items = [
            {"id": r["ID"], "nome": r["NomeSubUnidade"]} 
            for r in rows if r["NomeSubUnidade"]
        ]
        
        return jsonify({"items": items})
    except Exception as e:
        print(f"Erro ao buscar subseções: {e}")
        traceback.print_exc()
        return jsonify({"error": "Erro ao buscar subseções", "items": []}), 500


# -------------------------------------------------------
#            ENDPOINT PRINCIPAL: LISTA SIMPLES
# -------------------------------------------------------
# Correção do endpoint lista_simples no arquivo reports.py
# Substituir a função lista_simples existente por esta versão corrigida:

@bp.get("/lista_simples")
@require_auth
def lista_simples():
    try:
        # CORREÇÃO: Aceitar apenas formatos válidos
        formato = (request.args.get("formato") or "pdf").lower()
        subsecao = (request.args.get("subsecao") or "").strip()
        modo = (request.args.get("modo") or "").lower()  # "multi" => zip por subseção
        
        # NOVO: Receber campos selecionados e orientação
        campos_param = request.args.get("campos", "")
        campos_selecionados = [c.strip() for c in campos_param.split(",") if c.strip()] if campos_param else []
        orientacao = request.args.get("orientacao", "paisagem")  # padrão: paisagem
        
        # Debug: Log dos parâmetros recebidos
        print(f"DEBUG - Parâmetros recebidos:")
        print(f"  - formato: {formato}")
        print(f"  - subsecao: {subsecao}")
        print(f"  - orientacao: {orientacao}")
        print(f"  - campos_selecionados: {campos_selecionados}")

        # CORREÇÃO: Validação do formato - aceitar apenas pdf, xlsx, csv
        if formato not in ["pdf", "xlsx", "csv"]:
            return jsonify({
                "error": f"Formato inválido '{formato}'. Use: pdf, xlsx ou csv"
            }), 400

        # CORREÇÃO: Validação da orientação para PDFs
        if formato == "pdf" and orientacao not in ["retrato", "paisagem"]:
            print(f"AVISO: Orientação '{orientacao}' inválida, usando 'paisagem' como padrão")
            orientacao = "paisagem"

        # Busca os dados
        if subsecao:
            df = _consulta_lista_simples(subsecao)
            escopo = subsecao
        else:
            df = _consulta_lista_simples(None)
            escopo = "Geral"

        # NOVO: Filtrar colunas se campos foram especificados
        if campos_selecionados and not df.empty:
            # Mapear nomes do frontend para nomes do DataFrame
            campo_map = {
                "OAB": "OAB",
                "Nome": "Nome", 
                "CPF/CNPJ": "CPFCNPJ",
                "Situacao": "Situacao",
                "DataNascimento": "DataNascimento", 
                "DataCompromisso": "DataCompromisso",
                "TelefoneCelular": "TelefoneCelular",
                "Email": "Email",
                "Subsecao": "Subsecao"
            }
            
            # Converter campos selecionados para nomes de colunas do DataFrame
            colunas_filtradas = []
            for campo in campos_selecionados:
                if campo in campo_map and campo_map[campo] in df.columns:
                    colunas_filtradas.append(campo_map[campo])
            
            # Filtrar DataFrame apenas com as colunas selecionadas
            if colunas_filtradas:
                df = df[colunas_filtradas]

        # ---- PDF ----
        if formato == "pdf":
            # Quando geral + modo=multi => gera 1 PDF por subseção dentro de um ZIP
            if not subsecao and modo == "multi" and not df.empty and "Subsecao" in df.columns:
                subs = sorted([s for s in df["Subsecao"].dropna().unique().tolist() if s])
                if not subs:
                    return jsonify({"error": "Nenhuma subseção encontrada"}), 404
                    
                memzip = io.BytesIO()
                with zipfile.ZipFile(memzip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                    for s in subs:
                        dfi = df[df["Subsecao"] == s].reset_index(drop=True)
                        pdf = _pdf_from_df(dfi, "Relatório simples de Inscritos", s, campos_selecionados, orientacao)
                        # Nome de arquivo seguro (remove caracteres especiais)
                        safe_name = "".join(c for c in s if c.isalnum() or c in (' ', '-', '_')).rstrip()
                        zf.writestr(f"Relatorio_Lista_Simples_{safe_name}.pdf", pdf.getvalue())
                memzip.seek(0)
                return send_file(
                    memzip,
                    mimetype="application/zip",
                    as_attachment=True,
                    download_name="Relatorio_Lista_Simples_por_Subsecao.zip",
                )

            pdf = _pdf_from_df(df, "Relatório simples de Inscritos", escopo, campos_selecionados, orientacao)
            
            # CORREÇÃO: Nome do arquivo deve incluir orientação para melhor identificação
            orientacao_suffix = f"_{orientacao}" if orientacao == "retrato" else ""
            return send_file(
                pdf,
                mimetype="application/pdf",
                as_attachment=True,
                download_name=f"Relatorio_Lista_Simples_{escopo}{orientacao_suffix}.pdf",
            )

        # ---- XLSX ----
        elif formato == "xlsx":
            excel_file = _excel_from_df(df, "Relatório simples de Inscritos", escopo, campos_selecionados)
            return send_file(
                excel_file,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=f"Relatorio_Lista_Simples_{escopo}.xlsx",
            )

        # ---- CSV ----
        elif formato == "csv":
            csv_file = _csv_from_df(df, "Relatório simples de Inscritos", escopo, campos_selecionados)
            return send_file(
                csv_file,
                mimetype="text/csv; charset=utf-8",
                as_attachment=True,
                download_name=f"Relatorio_Lista_Simples_{escopo}.csv",
            )

    except Exception as e:
        print(f"Erro no endpoint lista_simples: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Erro interno do servidor: {str(e)}"}), 500

# -------------------------------------------------------
#                ENDPOINT DE TESTE
# -------------------------------------------------------
@bp.get("/test")
@require_auth
def test_endpoint():
    """Endpoint de teste para verificar se a API está funcionando."""
    return jsonify({
        "status": "ok",
        "message": "API de relatórios funcionando",
        "timestamp": dt.datetime.now().isoformat(),
        "user": request.user.get("name", "Desconhecido")
    })